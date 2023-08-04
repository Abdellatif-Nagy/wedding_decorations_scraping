from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.chrome.options import Options
from openpyxl import Workbook
from openpyxl.styles import NamedStyle,Font,Alignment, Color, Side, Border
import sys
######### creating empty lists

names = []
locations = []
links = []

######### main driver to get  names and locations
def main_driver():
    options = Options()
    options.add_argument("--headless=new")
    driver = webdriver.Chrome(options = options)
    url = f'https://www.easyweddings.com.au/WeddingDecorations/'
    driver.get(url)
    while True:
        decorations_names = driver.find_elements('xpath','(//div[@class="supplier-group-grid"]/div)/div/div[2]/span/h2')
        decorations_locations = driver.find_elements('xpath','//span[@class="card-address text-uppercase fw-bold"]')
        for n in decorations_names:
            names.append(n.text)
        for l in decorations_locations:
            locations.append(l.text)
        try:
            element = driver.find_element('xpath', '//a/i[@class="fa-light fa-chevron-right text-ew-lavendar"]')
            driver.execute_script("arguments[0].click();", element)
        except NoSuchElementException:
            print('pages are finished')
            break


    print(f'----> {len(names)} wedding decorations are founded')

    driver.quit()
    return driver

########## get the wedding decorations website
def get_website_link(names, loctions):
    options = Options()
    options.add_argument("--headless=new")
    options.binary_location = "C:/Program Files/Google/Chrome Dev/Application/chrome.exe"
    driver2 = webdriver.Chrome(options=options)
    driver2.get('https://www.google.com/search?q=hi&sxsrf=AB5stBiWdvYrE11OqQl3oQA1jfmtARsSRA%3A1691055574017&ei=1nXLZOU937uR1Q_cm4mQDw&ved=0ahUKEwjlgKP2mMCAAxXfXaQEHdxNAvIQ4dUDCA8&uact=5&oq=hi&gs_lp=Egxnd3Mtd2l6LXNlcnAiAmhpMg0QLhiKBRgKGAEYQxgqMgcQABiKBRhDMgcQABiKBRhDMgcQABiKBRhDMgUQABiABDIFEAAYgAQyCxAuGIAEGMcBGNEDMgUQABiABDILEC4YgAQYxwEY0QMyBRAuGIAEMhwQLhiKBRgKGAEYQxgqGJcFGNwEGN4EGOAE2AECSMm-FFDj5BNYs7sUcAN4ApABAJgBvwGgAfECqgEDMC4yuAEDyAEA-AEBqAILwgIEEAAYR8ICChAAGEcY1gQYsAPCAgoQABiKBRiwAxhDwgIHECMY6gIYJ8ICEBAAGIoFGOoCGLQCGEPYAQHCAgcQLhiKBRhDwgIIEAAYgAQYsQPCAgsQABiABBixAxiDAcICCxAuGIMBGLEDGIAE4gMEGAAgQYgGAZAGCboGBggBEAEYAboGBggCEAEYFA&sclient=gws-wiz-serp')
    print("Loading the links:")
    for i in range(len(names)):
        driver2.find_element('xpath', '//*[@id="tsf"]/div[1]/div[1]/div[2]/div/div[3]/div[1]').click()
        search_text = f'{names[i]} {loctions[i]} au'
        driver2.find_element('xpath','//*[@id="APjFqb"]').send_keys(f'{search_text}')
        search_button = driver2.find_element('xpath', '//*[@id="tsf"]/div[1]/div[1]/div[2]/button')
        search_button.click()
        link = driver2.find_element('xpath','//*[@id="rso"]/div[1]/div/div/div[1]/div/div/a')
        link = link.get_attribute('href')
        links.append(link)
        animation = ["[■□□□□□□□□□]", "[■■□□□□□□□□]", "[■■■□□□□□□□]", "[■■■■□□□□□□]", "[■■■■■□□□□□]", "[■■■■■■□□□□]","[■■■■■■■□□□]", "[■■■■■■■■□□]", "[■■■■■■■■■□]", "[■■■■■■■■■■]"]
        sys.stdout.write("\r" + animation[i % len(animation)])
        sys.stdout.flush()
    print("\n")
    return driver2
main_driver()
get_website_link(names, locations)

# creating the spreadsheet
wb = Workbook()
sheet = wb.active
sheet["A1"] = 'name'
sheet['B1'] = 'location'
sheet['C1'] = 'website'
# Let's create a style template for the header row
header = NamedStyle(name="header")
header.font = Font(bold=True)
header.border = Border(bottom=Side(border_style="thin"))
header.alignment = Alignment(horizontal="center", vertical="center")

# Now let's apply this to all first row (header) cells
header_row = sheet[1]
for cell in header_row:
    cell.style = header
####adding names , loctions, websites into the spreadsheet

#adding names
names_row_num = 2
names_col_num = 1
for n, value in enumerate(names, start= names_row_num):
    sheet.cell(row= n, column= names_col_num).value = value
#adding locations
locations_row_num = 2
locations_col_num = 2
for l, value in enumerate(locations, start= locations_row_num):
    sheet.cell(row= l, column= locations_col_num).value = value
#adding website for each name
site_row_num = 2
site_col_num = 3
for s, value in enumerate(links, start= site_row_num):
    sheet.cell(row= s, column= site_col_num).value = value

wb.save(filename= 'wedding_decorations.xlsx')
print("excel file named: wedding_decorations has been created")
